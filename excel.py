from openpyxl import Workbook, load_workbook
import os
answers = []


def save_to_excel():
    filename = "conversatii.xlsx"
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Conversatii"
    for answer in answers:
        ws.append([answer])
    wb.save("conversatii.xlsx")
    wb.close()
